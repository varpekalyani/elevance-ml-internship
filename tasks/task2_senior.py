"""
tasks/task2_senior.py
Task 2: Senior Citizen Identification via Video / Webcam
- Detects multiple persons
- Age > 60 → SENIOR CITIZEN (orange box)
- Logs age, gender, time → CSV + Excel
"""

import tkinter as tk
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
import cv2, numpy as np, threading, time, os, csv
from datetime import datetime
import pandas as pd
from tasks.shared import (BG, SURFACE, CARD, TEXT, MUTED, BORDER,
                           SENIOR, MALE_C, FEMALE_C,
                           ensure_weights, load_face_net,
                           load_age_net, load_gender_net,
                           detect_faces, predict_age_gender,
                           make_btn, card_frame)

ACCENT2  = "#FF6B35"
LOG_DIR  = os.path.join(os.path.dirname(os.path.dirname(__file__)), "logs")
CSV_PATH = os.path.join(LOG_DIR, "senior_log.csv")
HEADERS  = ["Timestamp", "Age", "Gender", "Senior Citizen", "Confidence"]


def init_log():
    os.makedirs(LOG_DIR, exist_ok=True)
    if not os.path.exists(CSV_PATH):
        with open(CSV_PATH, "w", newline="") as f:
            csv.writer(f).writerow(HEADERS)


def append_log(detections):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(CSV_PATH, "a", newline="") as f:
        w = csv.writer(f)
        for age, gender, conf, is_senior in detections:
            w.writerow([ts, age, gender,
                        "Yes" if is_senior else "No",
                        f"{conf:.0%}"])


def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        df = pd.read_csv(CSV_PATH)
        xls = CSV_PATH.replace(".csv", ".xlsx")
        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "Detection Log"
        thin = Border(
            left=Side(style="thin", color="333333"),
            right=Side(style="thin", color="333333"),
            top=Side(style="thin", color="333333"),
            bottom=Side(style="thin", color="333333"))
        for ci, h in enumerate(HEADERS, 1):
            c = ws.cell(1, ci, h)
            c.font      = Font(bold=True, color="C8FF00", name="Courier New")
            c.fill      = PatternFill("solid", fgColor="0D0D0D")
            c.alignment = Alignment(horizontal="center")
            c.border    = thin
        widths = [22, 8, 10, 16, 14]
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[
                ws.cell(1, ci).column_letter].width = w
        for ri, row in enumerate(df.itertuples(index=False), 2):
            senior = str(row[3]).strip().lower() == "yes"
            for ci, val in enumerate(row, 1):
                c = ws.cell(ri, ci, val)
                c.border    = thin
                c.alignment = Alignment(horizontal="center")
                c.fill      = PatternFill("solid",
                              fgColor="2D1A0D" if senior else "0D1A0D")
                c.font      = Font(color="FF6B35" if senior else "C8FF00",
                                   name="Courier New")
        wb.save(xls)
        return xls
    except Exception as e:
        return str(e)


class Task2Frame(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=BG)
        self._running   = False
        self._cap       = None
        self._face_net  = self._age_net = self._gen_net = None
        self._photo_ref = None
        self._stats     = {"total": 0, "senior": 0, "male": 0, "female": 0}
        init_log()
        self._build()
        threading.Thread(target=self._load_models, daemon=True).start()

    def _load_models(self):
        ensure_weights()
        self._face_net = load_face_net()
        self._age_net  = load_age_net()
        self._gen_net  = load_gender_net()
        self.after(0, lambda: self._status.configure(
            text="Models READY", fg=ACCENT2))

    def _build(self):
        hdr = tk.Frame(self, bg=BG)
        hdr.pack(fill="x", padx=16, pady=(12, 6))
        tk.Label(hdr, text="TASK 2  ·  Senior Citizen Identification",
                 font=("Courier", 13, "bold"), fg=ACCENT2, bg=BG).pack(side="left")
        self._status = tk.Label(hdr, text="Loading models…",
                                font=("Courier", 9), fg=MUTED, bg=BG)
        self._status.pack(side="right")

        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True, padx=16, pady=4)

        # Left — video canvas
        left = tk.Frame(body, bg=BG)
        left.pack(side="left", fill="both", expand=True, padx=(0, 12))

        self._canvas = tk.Canvas(left, bg=SURFACE, bd=0,
                                  highlightthickness=1,
                                  highlightbackground=BORDER)
        self._canvas.pack(fill="both", expand=True)
        self._draw_idle()

        btn_row = tk.Frame(left, bg=BG)
        btn_row.pack(fill="x", pady=(8, 0))
        self._btn_cam = make_btn(btn_row, "Start Webcam",
                                  self._start_webcam, primary=True, accent=ACCENT2)
        self._btn_cam.pack(side="left", expand=True, fill="x", padx=(0,6))
        self._btn_vid = make_btn(btn_row, "Load Video", self._load_video)
        self._btn_vid.pack(side="left", expand=True, fill="x", padx=(0,6))
        self._btn_stop = make_btn(btn_row, "Stop", self._stop, danger=True)
        self._btn_stop.pack(side="left", expand=True, fill="x")
        self._btn_stop.configure(state="disabled")

        exp_row = tk.Frame(left, bg=BG)
        exp_row.pack(fill="x", pady=(6, 0))
        make_btn(exp_row, "Open CSV", self._open_csv).pack(
            side="left", expand=True, fill="x", padx=(0,6))
        make_btn(exp_row, "Export Excel", self._do_excel).pack(
            side="left", expand=True, fill="x")

        # Right — stats + log
        right = tk.Frame(body, bg=BG, width=300)
        right.pack(side="right", fill="y")
        right.pack_propagate(False)

        tk.Label(right, text="LIVE STATS", font=("Courier", 9),
                 fg=MUTED, bg=BG).pack(anchor="w")
        tk.Frame(right, bg=BORDER, height=1).pack(fill="x", pady=(2, 8))

        stats_card = card_frame(right)
        stats_card.pack(fill="x")
        self._stat_lbls = {}
        for key, label, color in [
            ("total",  "Total detected",  TEXT),
            ("senior", "Senior (>60)",    ACCENT2),
            ("male",   "Male",            MALE_C),
            ("female", "Female",          FEMALE_C),
        ]:
            row = tk.Frame(stats_card, bg=CARD)
            row.pack(fill="x", padx=12, pady=4)
            tk.Label(row, text=label, font=("Courier", 9),
                     fg=MUTED, bg=CARD).pack(side="left")
            lbl = tk.Label(row, text="0",
                           font=("Courier", 15, "bold"),
                           fg=color, bg=CARD)
            lbl.pack(side="right")
            self._stat_lbls[key] = lbl

        tk.Label(right, text="LOG", font=("Courier", 9),
                 fg=MUTED, bg=BG).pack(anchor="w", pady=(16, 0))
        tk.Frame(right, bg=BORDER, height=1).pack(fill="x", pady=(2, 6))

        lf = tk.Frame(right, bg=BG)
        lf.pack(fill="both", expand=True)
        sb = tk.Scrollbar(lf, bg=BG, troughcolor=SURFACE)
        sb.pack(side="right", fill="y")
        self._log_box = tk.Listbox(lf, bg=SURFACE, fg=TEXT,
                                    font=("Courier", 8), bd=0,
                                    highlightthickness=0,
                                    yscrollcommand=sb.set)
        self._log_box.pack(fill="both", expand=True)
        sb.config(command=self._log_box.yview)

    def _draw_idle(self):
        self._canvas.delete("all")
        self._canvas.update_idletasks()
        w = self._canvas.winfo_width() or 600
        h = self._canvas.winfo_height() or 420
        self._canvas.create_text(w//2, h//2,
                                  text="Start Webcam or Load Video",
                                  fill=MUTED, font=("Courier", 12))

    def _start_webcam(self): self._start_capture(0)

    def _load_video(self):
        p = filedialog.askopenfilename(
            filetypes=[("Video", "*.mp4 *.avi *.mov *.mkv")])
        if p:
            self._start_capture(p)

    def _start_capture(self, source):
        if self._running:
            self._stop()
            time.sleep(0.3)
        self._cap = cv2.VideoCapture(source)
        if not self._cap.isOpened():
            messagebox.showerror("Error", f"Cannot open: {source}")
            return
        self._running = True
        self._btn_cam.configure(state="disabled")
        self._btn_vid.configure(state="disabled")
        self._btn_stop.configure(state="normal")
        threading.Thread(target=self._loop, daemon=True).start()

    def _stop(self):
        self._running = False
        if self._cap:
            self._cap.release()
            self._cap = None
        self._btn_cam.configure(state="normal")
        self._btn_vid.configure(state="normal")
        self._btn_stop.configure(state="disabled")
        self.after(200, self._draw_idle)

    def _loop(self):
        skip = 0
        while self._running:
            if not self._cap:
                break
            ret, frame = self._cap.read()
            if not ret:
                self.after(0, self._stop)
                break
            skip += 1
            detections = []
            if skip % 3 == 0 and self._face_net:
                boxes = detect_faces(frame, self._face_net)
                results = []
                for (x1, y1, x2, y2) in boxes:
                    face = frame[y1:y2, x1:x2]
                    if face.size == 0:
                        continue
                    age, gender, conf = predict_age_gender(
                        face, self._age_net, self._gen_net)
                    is_senior = age > 60
                    results.append((age, gender, conf, is_senior))
                    detections.append((age, gender, conf, is_senior))
                    color = (50, 107, 255) if is_senior else (0, 200, 100)
                    cv2.rectangle(frame, (x1,y1), (x2,y2), color, 2)
                    cv2.putText(frame, f"{gender},{age}",
                                (x1, max(y1-8,14)),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)
                    if is_senior:
                        cv2.putText(frame, "SENIOR",
                                    (x1, y2+20),
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.55, color, 2)
                if detections:
                    append_log(detections)
                    self.after(0, lambda d=detections: self._update_stats(d))

            self._push_frame(frame)

    def _push_frame(self, frame):
        self._canvas.update_idletasks()
        cw = max(self._canvas.winfo_width(), 500)
        ch = max(self._canvas.winfo_height(), 380)
        rgb   = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        img   = Image.fromarray(rgb)
        img.thumbnail((cw, ch), Image.LANCZOS)
        photo = ImageTk.PhotoImage(img)
        self._photo_ref = photo
        self.after(0, lambda p=photo: self._show(p))

    def _show(self, photo):
        self._canvas.delete("all")
        w = self._canvas.winfo_width()
        h = self._canvas.winfo_height()
        self._canvas.create_image(w//2, h//2, anchor="center", image=photo)

    def _update_stats(self, detections):
        for age, gender, conf, is_senior in detections:
            self._stats["total"] += 1
            if is_senior:
                self._stats["senior"] += 1
            if gender == "Male":
                self._stats["male"] += 1
            else:
                self._stats["female"] += 1
        for k, lbl in self._stat_lbls.items():
            lbl.configure(text=str(self._stats[k]))
        ts = datetime.now().strftime("%H:%M:%S")
        for age, gender, conf, is_senior in detections:
            entry = f"{ts}  {age}yrs  {gender}  {'SENIOR' if is_senior else ''}"
            self._log_box.insert(0, entry)

    def _open_csv(self):
        if os.path.exists(CSV_PATH):
            os.startfile(CSV_PATH) if os.name == "nt" else \
                os.system(f"xdg-open '{CSV_PATH}'")
        else:
            messagebox.showinfo("CSV", "No log yet.")

    def _do_excel(self):
        if not os.path.exists(CSV_PATH):
            messagebox.showinfo("Excel", "No log yet.")
            return
        result = export_excel()
        if result.endswith(".xlsx"):
            messagebox.showinfo("Excel", f"Saved → {result}")
        else:
            messagebox.showerror("Excel", result)

    def on_close(self):
        self._stop()
